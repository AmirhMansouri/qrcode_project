import qrcode
import openpyxl
import PIL
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
loc = "file1.xlsx"
wb_obj = openpyxl.load_workbook(loc)
sheet_obj = wb_obj.active
for i in range(2, sheet_obj.max_row+1):
    myFont = ImageFont.truetype( 'arial.ttf' ,35)
    data = sheet_obj.cell(row = i, column = 1)
    img = qrcode.make(data.value)
    NAMe = sheet_obj.cell(row = i, column = 2)
    img.save(str(NAMe.value)+".png")
    img = Image.open(str(NAMe.value)+".png")
    I1 = ImageDraw.Draw(img)
    I1.text((75, 7), str(NAMe.value),font = myFont)
    img.save(str(NAMe.value)+".png")